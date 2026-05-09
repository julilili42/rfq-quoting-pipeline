"""Deterministic fast-path for trivial RFQs.

When a customer mail (or attachment) literally names a known EK article
number plus a clearly-stated quantity, we don't need an LLM call to
extract that. This module:

1. Builds an Aho-Corasick automaton over all stammdaten article numbers
   once at startup.
2. On each new mail, normalises the text + every text-bearing attachment
   and scans for known article numbers.
3. Looks for a quantity in an 80-char window *after* each hit, using a
   small set of regex patterns covering English + German.
4. If exactly one quantity is found per unique article number, returns
   a fully-populated ``Anfrage`` (description/material/dims come from
   stammdaten, einheit defaults to ``Stk``). Otherwise returns ``None``
   and the pipeline falls back to the LLM.

This is a sieve, not a parser. Any ambiguity → fall through. Designed to
be wrong about *whether* to fire, never wrong about *what* the answer is
when it does fire.
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from pathlib import Path

from ..core import Anfrage, Position
from ..data import StammdatenRecord, StammdatenRepository
from ..ingestion import Mail

log = logging.getLogger(__name__)

# Window in original-text chars searched after each article-nr hit.
QUANTITY_WINDOW = 80

# Article-nr patterns shorter than this are too noisy to safely match.
MIN_ARTIKELNR_LEN = 6

# Quantity patterns ordered by specificity. First match wins per window.
# Operate on ORIGINAL text (not normalised), so spaces/case matter.
_QTY_PATTERNS = [
    # "qty 3", "qty: 3", "Menge: 3", "count 3"
    re.compile(r"\b(?:qty|quantity|menge|count)\b[\s:=]*?(\d{1,5})\b", re.IGNORECASE),
    # "6 pcs", "100 St.", "6 Stück", "10 ea", "10 pieces"
    re.compile(r"(\d{1,5})\s*(?:pcs?|stk\.?|st\.?|stueck|stück|pieces?|pc|ea)\b", re.IGNORECASE),
    # "x6", "× 100"
    re.compile(r"[xX×]\s*(\d{1,5})\b"),
]

# Dimensions like "55x63x6", "1.1x0.5x0.31", "55 x 63 x 6", "55,5x63,2x6,1"
# Stripped from window before quantity-pattern scan.
_DIMENSIONS = re.compile(
    r"\b\d+(?:[.,]\d+)?\s*[xX×]\s*\d+(?:[.,]\d+)?\s*[xX×]\s*\d+(?:[.,]\d+)?\b"
)


@dataclass(frozen=True)
class _Hit:
    """One Aho-Corasick hit, with positions in the original (un-normalised) text."""
    segment_idx: int       # index into the segment list
    orig_start: int        # inclusive
    orig_end: int          # inclusive
    record: StammdatenRecord


# ----------------------------------------------------------- normalisation

def _normalize_with_map(text: str) -> tuple[str, list[int]]:
    """Uppercase + strip ``[.\\-_ \\t\\n\\r]``. Returns (normalised, idx_map).

    ``idx_map[i]`` is the position in the original ``text`` of the i-th
    character in the normalised string. Used to recover the original
    span for each Aho-Corasick hit so the quantity window scans real
    user text (with spaces, lowercase, units…).
    """
    out_chars: list[str] = []
    out_idx: list[int] = []
    for i, c in enumerate(text):
        if c in ".-_ \t\n\r/":
            continue
        out_chars.append(c.upper())
        out_idx.append(i)
    return "".join(out_chars), out_idx


def _normalize(text: str) -> str:
    """Same normalisation, without the index map."""
    return _normalize_with_map(text)[0]


# ------------------------------------------------------ text collection

def _pdf_text(path: Path) -> str:
    """Plain text from a PDF via PyMuPDF (already a project dependency)."""
    try:
        import fitz  # PyMuPDF
    except ImportError:
        log.warning("PyMuPDF missing — fast-path skips PDF: %s", path.name)
        return ""
    parts: list[str] = []
    try:
        with fitz.open(path) as doc:
            for page in doc:
                parts.append(page.get_text())
    except Exception as exc:
        log.warning("Fast-path could not read PDF %s: %s", path.name, exc)
        return ""
    return "\n".join(parts)


def _xlsx_text(path: Path) -> str:
    """Flatten every sheet's cell values to a single text blob."""
    try:
        import openpyxl
    except ImportError:
        log.warning("openpyxl missing — fast-path skips XLSX: %s", path.name)
        return ""
    parts: list[str] = []
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        log.warning("Fast-path could not read XLSX %s: %s", path.name, exc)
        return ""
    try:
        for sn in wb.sheetnames:
            ws = wb[sn]
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None and str(c).strip()]
                if cells:
                    parts.append(" ".join(cells))
    finally:
        wb.close()
    return "\n".join(parts)


def _collect_segments(mail: Mail) -> list[tuple[str, str]]:
    """Return ``[(source_label, text), …]`` — body + every text-bearing attachment.

    ``source_label`` is what we put in :attr:`Position.source_file`:
    ``"mail"`` for the body, the file name otherwise.
    """
    segments: list[tuple[str, str]] = []
    if mail.body:
        segments.append(("mail", mail.body))
    for att in mail.attachments:
        ext = att.suffix.lower()
        if ext == ".pdf":
            text = _pdf_text(att)
            if text:
                segments.append((att.name, text))
        elif ext in (".xlsx", ".xls"):
            text = _xlsx_text(att)
            if text:
                segments.append((att.name, text))
        # Other types (images, .gif, .csv, .docx) skipped — fall through to LLM.
    return segments


# ------------------------------------------------------ quantity extraction

def _extract_quantity(window: str) -> int | None:
    """Find a plausible quantity in ``window``. Strips dimensions first."""
    cleaned = _DIMENSIONS.sub(" ", window)
    for pattern in _QTY_PATTERNS:
        m = pattern.search(cleaned)
        if m:
            try:
                qty = int(m.group(1))
            except ValueError:
                continue
            if 1 <= qty <= 99999:
                return qty
    return None


# ------------------------------------------------------ public API

class FastPathExtractor:
    """Deterministic Anfrage-builder for trivial RFQs.

    Construct once with the stammdaten repository. The Aho-Corasick
    automaton is built lazily on first call so empty-stammdaten test
    paths don't pay the cost.
    """

    def __init__(self, repo: StammdatenRepository):
        self._repo = repo
        self._automaton = None  # ahocorasick.Automaton

    def _ensure_built(self) -> None:
        if self._automaton is not None:
            return
        import ahocorasick
        a = ahocorasick.Automaton()
        added = 0
        for rec in self._repo.all():
            norm = _normalize(rec.artikel_nr)
            if len(norm) < MIN_ARTIKELNR_LEN:
                continue
            a.add_word(norm, (norm, rec))
            added += 1
        if added > 0:
            a.make_automaton()
        self._automaton = a
        log.info("Fast-path automaton built: %d article numbers", added)

    def try_extract(self, mail: Mail) -> Anfrage | None:
        """Return a fully-populated ``Anfrage`` or ``None`` to fall through."""
        self._ensure_built()
        if self._automaton is None or len(self._automaton) == 0:
            return None

        segments = _collect_segments(mail)
        if not segments:
            return None

        hits = self._scan(segments)
        if not hits:
            return None

        # Hits are scanned per segment in document order, but `iter`
        # returns matches in trie-traversal order, so re-sort to be safe.
        hits.sort(key=lambda h: (h.segment_idx, h.orig_start))

        # Keep the first occurrence of each unique article number — but
        # remember the position of the *next* article hit (any article)
        # so the quantity window for hit[i] never bleeds into hit[i+1].
        first_per_artnr: list[_Hit] = []
        seen: set[str] = set()
        for h in hits:
            if h.record.artikel_nr in seen:
                continue
            seen.add(h.record.artikel_nr)
            first_per_artnr.append(h)

        positions: list[Position] = []
        for pos_idx, hit in enumerate(first_per_artnr, start=1):
            seg_label, seg_text = segments[hit.segment_idx]
            # Window: from end-of-this-hit up to either +QUANTITY_WINDOW
            # chars or the next any-article hit in this segment.
            window_start = hit.orig_end + 1
            window_end = window_start + QUANTITY_WINDOW
            for next_hit in hits:
                if (
                    next_hit.segment_idx == hit.segment_idx
                    and next_hit.orig_start > hit.orig_end
                    and next_hit.orig_start < window_end
                ):
                    window_end = next_hit.orig_start
                    break
            window = seg_text[window_start:window_end]
            qty = _extract_quantity(window)
            if qty is None:
                # One missing quantity = ambiguous → bail out, let LLM handle.
                log.info("Fast-path declines: no quantity for %s", hit.record.artikel_nr)
                return None

            positions.append(_position_from_record(
                pos_nr=pos_idx * 10,
                record=hit.record,
                menge=qty,
                source_file=seg_label,
                source_quote=_extract_source_quote(seg_text, hit),
            ))

        anfrage = Anfrage(
            kunde_firma=_sender_company(mail.sender),
            kunde_email=_sender_email(mail.sender),
            positionen=positions,
            unsicherheiten=[],
        )
        log.info(
            "Fast-path matched %d position(s): %s",
            len(positions),
            ", ".join(p.artikelnummer for p in positions),
        )
        return anfrage

    def _scan(self, segments: list[tuple[str, str]]) -> list[_Hit]:
        out: list[_Hit] = []
        for seg_idx, (_label, text) in enumerate(segments):
            norm, idx_map = _normalize_with_map(text)
            for end_pos, (pattern, rec) in self._automaton.iter(norm):
                start_pos = end_pos - len(pattern) + 1
                if start_pos < 0 or end_pos >= len(idx_map):
                    continue
                out.append(_Hit(
                    segment_idx=seg_idx,
                    orig_start=idx_map[start_pos],
                    orig_end=idx_map[end_pos],
                    record=rec,
                ))
        return out


# ------------------------------------------------------ position builder

def _position_from_record(
    *,
    pos_nr: int,
    record: StammdatenRecord,
    menge: int,
    source_file: str,
    source_quote: str,
) -> Position:
    return Position(
        pos_nr=pos_nr,
        artikelnummer=record.artikel_nr,
        bezeichnung=record.bezeichnung or "",
        menge=float(menge),
        einheit=record.einheit or "Stk",
        werkstoff=record.werkstoff,
        abmessungen=record.abmessungen,
        confidence="high",
        source_quote=source_quote,
        source_file=source_file,
    )


def _extract_source_quote(text: str, hit: _Hit, padding: int = 40) -> str:
    """Return a snippet of the original text around the hit, for the UI."""
    start = max(0, hit.orig_start - padding)
    end = min(len(text), hit.orig_end + 1 + padding)
    snippet = text[start:end]
    return " ".join(snippet.split())  # collapse whitespace for display


# Sender parsing for "Name <email@host>" format. Forwarded mail bodies
# often clobber the From header; this is a best-effort header read only.
_SENDER_EMAIL = re.compile(r"<([^>]+@[^>]+)>")


def _sender_email(sender: str) -> str | None:
    if not sender:
        return None
    m = _SENDER_EMAIL.search(sender)
    if m:
        return m.group(1).strip()
    if "@" in sender:
        return sender.strip()
    return None


def _sender_company(sender: str) -> str | None:
    if not sender:
        return None
    name = _SENDER_EMAIL.sub("", sender).strip().strip('"').strip()
    return name or None
